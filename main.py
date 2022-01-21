import os
import csv
import datetime
import requests
from requests.auth import HTTPBasicAuth
import json
from collections import OrderedDict
from openpyxl import load_workbook

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and shttp://192.168.1.67:8080/wallbrain-ws/api/firewall/all/listettings.
#print(os.getcwd())

if not os.path.isfile("properties.ini"):
    print("properties.ini 생성")
    f= open('properties.ini', 'w')
    f.write("http://192.168.1.67:8080/wallbrain-ws/api;root;!q1w2e3r4")

if not os.path.isfile("FPMS_test.csv"):
    print("FPMS_test.csv 생성")
    f = open(log_file, "w", encoding='utf-8', newline='')
    f.writelines("[구분],[Src IP],[Dst IP],[Dst Protocol],[Dst Port],[Compl 예상결과],,[Discover 예상결과],\r\n")


file = open("properties.ini")
parameters = file.read().split(";")


URL = parameters[0]
USER = parameters[1]
PW = parameters[2]
basic = HTTPBasicAuth(USER, PW)

print(parameters)
#exit()
BASE_HEADERS = {
    "Connection":"keep-alive",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*",
    "Content-Type":"application/json; charset=UTF-8"
    }
def get_status_code(resultcode):
    print(resultcode)
def get_firewall_all_list():
    response = requests.get(URL+'/firewall/all/list',headers=BASE_HEADERS,auth=basic)
    if response.status_code == 200:
        result = json.loads(response.content)

        return result
    else:
        get_status_code(response.status_code)
        return response.status_code
def post_provision_rule_search(data):
    response = requests.post(URL+'/provision/rule/search',headers=BASE_HEADERS,json=data,auth=basic)
    if response.status_code == 200:
        print("200 OK")
        result = json.loads(response.content)
        return result

    else:
        get_status_code(response.status_code)
        print(str(response.status_code)+" NOT OK")
        data = {
            "error" : str(response.status_code)
        }
        return data

def get_now_datetime():
    str = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    return str

def get_csv_qa():
    start_time=get_now_datetime()
    total_cnt=0
    true_cnt=0
    false_cnt=0
    with open('FPMS_test.csv', newline='', encoding='utf-8') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        log_file=get_now_datetime()+"_FPMS_test_result.csv"
        f = open(log_file, "w",encoding='utf-8', newline='')
        f.writelines("구분,Src IP,Dst IP,Dst Protocol,Dst Port,Compl 예상결과,Compl 실제결과,Discover 예상결과,Discover 실제결과\r\n")
        for row in spamreader:
            total_cnt=total_cnt+1
            str = ', '.join(row)
            #print(str)
            if len(str.split(',')[1].split('.')) == 4:
                true_cnt = true_cnt + 1
                data = {
                    "srcAddr": str.split(',')[1],
                    "dstAddr": str.split(',')[2],
                    "dstPort": str.split(',')[4],
                    "protocol": str.split(',')[3],
                    "withDiscovery": "true",
                    "checkDuplicated": "false",
                    "duplicatedPriority": "false",
                    "sortByFirewallZone": "true"
                }
                post_provision_rule_search_result = post_provision_rule_search(data)
                print(str+','+post_provision_rule_search_result[0]['complianceComment'].replace(',','|')+','+','+'|'.join(post_provision_rule_search_result[0]['firewallIds']))
                #print(str.split(',')[5]+" | "+post_provision_rule_search_result[0]['complianceComment'])
                #print(post_provision_rule_search_result[0]['firewallIds'])
                csv_data=str+','+post_provision_rule_search_result[0]['complianceComment'].replace(',','|')+','+','+'|'.join(post_provision_rule_search_result[0]['firewallIds'])+"\r\n"
                f.writelines(csv_data)
            else:
                false_cnt=false_cnt+1
                print("error")
        f.close()
        print('start_time: '+start_time)
        print('end_time: ' + get_now_datetime())
        print('-----------------total_cnt-----------------')
        print(total_cnt)
        print('-----------------OK_cnt-----------------')
        print(true_cnt)
        print('-----------------Fail_cnt-----------------')
        print(false_cnt)
            #print(data)
            #post_provision_rule_search_result = post_provision_rule_search(data)
            #print(post_provision_rule_search_result)
            #print(post_provision_rule_search_result[0]['complianceComment'])
            #print(post_provision_rule_search_result[0]['firewallIds'])
            #print(row)
            #str=', '.join(row)
            #print(str.split(','))

def get_office_qa():
    print(os.getcwd())
    load_wb = load_workbook(filename='FPMS_test.csv', data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['FPMS']
    # 셀 주소로 값 출력
    print(load_ws['A1'].value)

def get_office():
    # data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook("FPMS_test.csv", data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['FPMS']
    # 셀 주소로 값 출력
    print(load_ws['A1'].value)
    # 셀 좌표로 값 출력
    print(load_ws.cell(1, 2).value)
    print('\n-----지정한 셀 출력-----')
    get_cells = load_ws['A1':'D2']
    for row in get_cells:
            for cell in row:
                print(cell.value)
    print('\n-----모든 행 단위로 출력-----')
    for row in load_ws.rows:
        print(row)
    print('\n-----모든 열 단위로 출력-----')
    for column in load_ws.columns:
        print(column)
    print('\n-----모든 행과 열 출력-----')
    all_values = []
    for row in load_ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)
    print(all_values)
    # 이름이 있는 시트를 생성
    # write_ws = write_wb.create_sheet('생성시트')

    # Sheet1에다 입력
    write_ws = write_wb.active
    write_ws['A1'] = '숫자'

    # 행 단위로 추가
    write_ws.append([1, 2, 3])

    # 셀 단위로 추가
    write_ws.cell(5, 5, '5행5열')
    write_wb.save('/Users/Jamong/Desktop/숫자.xlsx')

def writeToTextFile(a):
    fileWriter=open("test.txt",'w')
    fileWriter.write(str(a)+" is writed")
# Press the green button in the gutter to run the script.

if __name__ == '__main__':
    #print(basic)
    #response=get_firewall_all_list()
    #print(response[0]['firewallUuid'])
    #print("post_provision_rule_search")
    #data = {
    #    "srcAddr": "192.168.100.100",
    #    "dstAddr": "10.10.10.10",
    #    "dstPort": "45",
    #    "protocol": "TCP",
    #    "withDiscovery": "true",
    #    "checkDuplicated": "true"
    #}
    #post_provision_rule_search_result=post_provision_rule_search(data)
    #print(post_provision_rule_search_result)
    #print(post_provision_rule_search_result[0]['complianceComment'])
    #print(post_provision_rule_search_result[0]['firewallIds'])
    get_csv_qa()
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
